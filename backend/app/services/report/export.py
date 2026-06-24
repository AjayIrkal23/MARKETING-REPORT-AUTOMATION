"""Report JSW/JVML business logic: export the generated RAKE-pivot report as .xlsx."""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

from ...schemas.report import ReportQuery, ReportResponse
from ...utils.report.excel_style import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    add_premium_metadata_sheet,
    add_title_banner,
    apply_credit_balance_font,
    apply_pivot_view,
    apply_status_font,
    auto_size_columns,
    enable_table_filter,
    fmt_inr,
    fmt_quantity,
    freeze_header_and_fixed_cols,
    setup_print_page,
    style_data_row,
    style_grand_total_row,
    style_premium_header_row,
    style_subtotal_row,
)
from .generate import generate_report


# Fixed row fields in display order.
_FIXED_HEADERS = [
    "Distr. Channel",
    "Sold To Party",
    "BRANCH",
    "Party Code",
    "Ship To Party",
    "Transport Mode",
    "Destination",
    "ROUTE",
]

# Number of leading fixed columns (always shown, eligible for blanking).
_FIXED_COL_COUNT = 5

# Fixed trailing columns after the dynamic RAKE columns.
_TRAILING_HEADERS = [
    "Total",
    "Yes+DO",
    "Blocked",
    "Credit Balance",
    "Required Credit",
    "Credit Note",
]

# Optional-column key aligned to each header position (None ⇒ always shown). The
# export `columns` filter (the same keys the UI toggles) gates the non-None ones:
# the 5 fixed left cols and the dynamic RAKE cols are always shown; the 3 detail
# cols + the 6 trailing cols (Total included) are toggleable.
_FIXED_OPT_KEYS: list[str | None] = [None, None, None, None, None, "transport_mode", "destination", "route"]
_TRAILING_OPT_KEYS: list[str | None] = ["total", "yes_do", "blocked", "credit_balance", "required_credit", "credit_note"]

# Headers that carry numeric (quantity) formatting.
_QUANTITY_HEADERS = {"Total", "Yes+DO"}

# Headers that carry INR currency formatting.
_INR_HEADERS = {"Credit Balance", "Required Credit"}

# Headers whose text should be centred.
_CENTER_HEADERS = {"Blocked", "Credit Note"}


def _kept_indices(rake_n: int, visible: set[str] | None) -> list[int]:
    """Column indices to keep. visible=None ⇒ all columns; else always-on + visible keys."""
    keys = _FIXED_OPT_KEYS + [None] * rake_n + _TRAILING_OPT_KEYS
    if visible is None:
        return list(range(len(keys)))
    return [i for i, k in enumerate(keys) if k is None or k in visible]


def _sel(values: list[Any], kept: list[int]) -> list[Any]:
    return [values[i] for i in kept]


def _fmt_blocked(blocked: bool | None, credit_status: str) -> str:
    """Blocked cell text matching the frontend: Blocked / No / N/A."""
    if credit_status == "NO CREDIT REPORT FOUND":
        return "N/A"
    if blocked is None:
        return ""
    return "Blocked" if blocked else "No"


def _fmt_credit_balance(balance: float | None, credit_status: str) -> float | str:
    """Credit Balance value or status text when no report/balance exists."""
    if balance is not None:
        return balance
    if credit_status == "NO CREDIT REPORT FOUND":
        return "NO CREDIT REPORT FOUND"
    if credit_status == "No Credit balance":
        return "No Credit balance"
    return ""


def _fmt_num(value: float | None) -> float | str:
    return "" if value is None else value


# Leading fixed cols eligible for repeated-parent blanking (mirrors the frontend
# FIXED_COL_KEYS in report-grouping.ts). The 3 detail cols always show.
_BLANK_KEYS = ("distr_chnl", "sold_to_party", "sales_office", "party_code", "ship_to_party")


def _group_key(row: Any) -> tuple[str, str]:
    """Channel-group identity: SO Sales Org + Distr.Chnl (a label can recur across orgs)."""
    return (row.so_sales_org or "", row.distr_chnl or "")


def _blank_flags(row: Any, prev: Any | None) -> dict[str, bool]:
    """Blank a leading fixed cell while the whole parent chain above also matches."""
    flags: dict[str, bool] = {}
    parent_same = prev is not None
    for key in _BLANK_KEYS:
        same = parent_same and prev is not None and getattr(prev, key) == getattr(row, key)
        flags[key] = same
        parent_same = same
    return flags


def _empty_agg() -> dict[str, Any]:
    return {"rake": {}, "total": 0.0, "nco_yes_do": 0.0, "required_credit": None}


def _add_to_agg(agg: dict[str, Any], row: Any) -> None:
    for col, val in row.rake_quantities.items():
        agg["rake"][col] = agg["rake"].get(col, 0.0) + (val or 0.0)
    agg["total"] += row.total or 0.0
    agg["nco_yes_do"] += row.nco_yes_do or 0.0
    if row.required_credit is not None:
        agg["required_credit"] = (agg["required_credit"] or 0.0) + row.required_credit


def _fixed_cells(row: Any, flags: dict[str, bool]) -> list[Any]:
    """The 8 fixed/detail cells; blanked parents render empty, detail cols always show."""
    return [
        "" if flags["distr_chnl"] else (row.distr_chnl or ""),
        "" if flags["sold_to_party"] else (row.sold_to_party or ""),
        "" if flags["sales_office"] else (row.sales_office or ""),
        "" if flags["party_code"] else row.party_code,
        "" if flags["ship_to_party"] else (row.ship_to_party or ""),
        row.transport_mode or "",
        row.destination or "",
        row.route or "",
    ]


def _append_subtotal(
    ws,
    report: ReportResponse,
    agg: dict[str, Any],
    channel: str | None,
    kept: list[int],
) -> None:
    """Per-channel '{channel} Total' row: label in Distr.Channel col; RAKE/Total/Yes+DO/Required summed."""
    sub: list[Any] = [f"{channel or '—'} Total"] + [""] * (len(_FIXED_HEADERS) - 1)
    sub += [agg["rake"].get(col, 0.0) or "" for col in report.rake_columns]
    sub += [
        _fmt_num(agg["total"]),
        _fmt_num(agg["nco_yes_do"]),
        "",
        "",
        _fmt_num(agg["required_credit"]),
        "",
    ]
    ws.append(_sel(sub, kept))


def _write_report(
    ws, report: ReportResponse, visible: set[str] | None
) -> tuple[list[str], list[str]]:
    """Write the grouped pivot to *ws*: title banner + header + data rows (repeated
    parents blanked) + per-group Distr.Channel subtotal rows + a Grand Total.

    Returns the final list of header labels and the active RAKE column names so
    the styling pass knows which columns to format.
    """
    kept = _kept_indices(len(report.rake_columns), visible)
    headers = _FIXED_HEADERS + report.rake_columns + _TRAILING_HEADERS
    filtered_headers = _sel(headers, kept)
    n_cols = len(filtered_headers)

    # Title banner occupies rows 1-2; the table starts at row 3.
    report_label = "JSW" if report.report_type.lower() == "jsw" else "JVML"
    title = f"{report_label} Coil Stock Report"
    subtitle = (
        f"Report Date: {report.date}  |  Region: {report.region_name}  |  "
        f"Days Filter: {report.days_filter}  |  Exported: "
        f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
    )
    add_title_banner(ws, 1, n_cols, title, subtitle)

    # Header row.
    ws.append(filtered_headers)

    prev: Any | None = None
    agg = _empty_agg()

    for row in report.rows:
        if prev is not None and _group_key(prev) != _group_key(row):
            _append_subtotal(ws, report, agg, prev.distr_chnl, kept)
            agg = _empty_agg()
            prev = None  # new group → show the full parent chain again

        full = (
            _fixed_cells(row, _blank_flags(row, prev))
            + [row.rake_quantities.get(col, 0.0) or "" for col in report.rake_columns]
            + [
                _fmt_num(row.total),
                _fmt_num(row.nco_yes_do),
                _fmt_blocked(row.blocked, row.credit_status),
                _fmt_credit_balance(row.credit_balance, row.credit_status),
                _fmt_num(row.required_credit),
                row.credit_note,
            ]
        )
        ws.append(_sel(full, kept))
        _add_to_agg(agg, row)
        prev = row

    if prev is not None:
        _append_subtotal(ws, report, agg, prev.distr_chnl, kept)

    # Grand total row (credit money is intentionally NOT summed).
    total_row: list[Any] = ["Grand Total"] + [""] * (len(_FIXED_HEADERS) - 1)
    total_row += [""] * len(report.rake_columns)
    total_row += [
        _fmt_num(report.grand_total),
        _fmt_num(report.grand_nco_yes_do),
        "",
        "",
        _fmt_num(report.grand_required_credit),
        "",
    ]
    ws.append(_sel(total_row, kept))

    return filtered_headers, report.rake_columns


def _apply_column_formats(
    ws: Worksheet, row_idx: int, headers: list[str], rake_columns: list[str]
) -> None:
    """Apply numeric/INR formats, alignments and conditional font colours per column."""
    quantity_headers = _QUANTITY_HEADERS | set(rake_columns)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)

        if header in quantity_headers:
            cell.alignment = ALIGN_RIGHT
            fmt_quantity(cell)
            continue

        if header in _INR_HEADERS:
            cell.alignment = ALIGN_RIGHT
            if isinstance(cell.value, (int, float)):
                fmt_inr(cell)
            if header == "Credit Balance":
                apply_credit_balance_font(cell)
            continue

        if header == "Blocked":
            cell.alignment = ALIGN_CENTER
            apply_status_font(cell, str(cell.value))
            continue

        if header == "Credit Note":
            cell.alignment = ALIGN_CENTER
            apply_status_font(cell, str(cell.value))
            continue

        # Default text columns.
        cell.alignment = ALIGN_LEFT


def _style_report(ws: Worksheet, headers: list[str], rake_columns: list[str]) -> None:
    """Apply premium styling to the worksheet after all data has been written."""
    n_cols = len(headers)
    header_row = 3

    # Header.
    style_premium_header_row(ws, header_row, n_cols)

    # Body rows: identify by content and apply group alternation.
    group_idx = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        first_val = ws.cell(row=row_idx, column=1).value

        if first_val == "Grand Total":
            style_grand_total_row(ws, row_idx, n_cols, _FIXED_COL_COUNT)
            continue

        if isinstance(first_val, str) and first_val.endswith(" Total"):
            style_subtotal_row(ws, row_idx, n_cols, _FIXED_COL_COUNT)
            group_idx += 1
            continue

        is_alt_group = group_idx % 2 == 1
        style_data_row(ws, row_idx, n_cols, is_alt_group=is_alt_group)
        _apply_column_formats(ws, row_idx, headers, rake_columns)
        # Detail rows sit one outline level below their subtotal so each channel
        # group collapses/expands with Excel's native +/- buttons (pivot feel).
        ws.row_dimensions[row_idx].outline_level = 1

    # Column-specific alignment overrides for the header row.
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        if header in _QUANTITY_HEADERS or header in _INR_HEADERS or header in _CENTER_HEADERS:
            cell.alignment = ALIGN_CENTER if header in _CENTER_HEADERS else ALIGN_RIGHT
        else:
            cell.alignment = ALIGN_LEFT

    # Layout — pivot canvas (no gridlines, collapsible groups), header-only freeze
    # (no fixed/frozen columns), auto-filter and auto-sized columns.
    apply_pivot_view(ws)
    enable_table_filter(ws, header_row, n_cols, ws.max_row)
    freeze_header_and_fixed_cols(ws, header_row, 0)
    ws.row_dimensions[header_row].height = 26
    auto_size_columns(ws)


async def export_report(query: ReportQuery) -> bytes:
    """Generate the report and return it as a styled .xlsx workbook."""
    import openpyxl

    report = await generate_report(query)

    # None ⇒ all optional columns (old clients / direct hits); a CSV ⇒ only those keys.
    visible = None if query.columns is None else {c for c in query.columns.split(",") if c}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    headers, rake_columns = _write_report(ws, report, visible)
    _style_report(ws, headers, rake_columns)
    setup_print_page(ws, ws.title)

    add_premium_metadata_sheet(
        wb,
        title="Report Export",
        items={
            "Export time": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "Report date": query.date,
            "Report type": query.report_type.upper(),
            "Region": report.region_name,
            "Days filter": query.days,
            "CCAs": ", ".join(report.ccas),
            "Has stock": str(report.has_stock),
            "Has credit report": str(report.has_credit_report),
            "Coil price/qty": str(report.coil_price_per_qty) if report.coil_price_per_qty is not None else "Not set",
        },
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
