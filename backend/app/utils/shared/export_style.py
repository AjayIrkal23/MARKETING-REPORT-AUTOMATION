"""Shared Excel styling helpers for backend report exports.

All functions lazily import ``openpyxl`` so that importing modules that use
this util does not fail when the package is absent during development.
"""

from __future__ import annotations

from typing import Any


def style_header_row(row: tuple[Any, ...]) -> None:
    """Apply bold font, light-blue fill and thin borders to every cell in ``row``."""
    import openpyxl
    from openpyxl.styles import Border, Font, PatternFill, Side

    font = Font(bold=True)
    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in row:
        cell.font = font
        cell.fill = fill
        cell.border = border


def auto_size_columns(ws, min_width: int = 10, max_width: int = 60) -> None:
    """Resize every column to the widest populated cell, clamped to bounds."""
    import openpyxl

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        width = max(min_width, min(max_width, max_len + 4))
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width


def add_metadata_sheet(wb, title: str, items: dict[str, str]) -> None:
    """Append a ``Metadata`` sheet with a title and key/value rows."""
    ws = wb.create_sheet(title="Metadata")

    ws["A1"] = title
    format_report_title_cell(ws["A1"])

    ws["A3"] = "Key"
    ws["B3"] = "Value"
    style_header_row(ws[3])

    for row_idx, (key, value) in enumerate(items.items(), start=4):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=value)


def enable_filters(ws) -> None:
    """Enable auto-filter across the used range of ``ws``."""
    ws.auto_filter.ref = ws.dimensions


def format_report_title_cell(cell) -> None:
    """Style a report title cell with bold, larger font."""
    from openpyxl.styles import Font

    cell.font = Font(bold=True, size=14)
