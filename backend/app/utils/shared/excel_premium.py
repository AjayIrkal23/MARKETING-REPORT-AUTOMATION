"""Premium flat-table Excel writer shared by the stock / credit / customer-code
exports and the report rake-breakdown sheets.

Builds on the report styling primitives (``utils/report/excel_style``) so every
export shares one JSW-branded look: blue title banner, slate header, zebra rows,
branded number formats, frozen header, auto-filter and print-ready layout. This
is the single home for flat (non-grouped) sheets — domains pass headers + rows +
per-column format kinds and never re-implement styling.
"""

from __future__ import annotations

import re
from collections.abc import Iterable
from datetime import datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from ..report.excel_style import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    add_title_banner,
    apply_credit_balance_font,
    auto_size_columns,
    enable_table_filter,
    fmt_inr,
    fmt_quantity,
    freeze_header_and_fixed_cols,
    setup_print_page,
    style_data_row,
    style_grand_total_row,
    style_premium_header_row,
)

# Rows 1-2 are the banner; the column header sits on row 3 (matches the pivot).
HEADER_ROW = 3

# Excel forbids these chars in a sheet name and caps the name at 31 chars.
_ILLEGAL_SHEET = re.compile(r"[\[\]:*?/\\]")


def safe_sheet_name(base: str, used: set[str], *, suffix: str = "") -> str:
    """Return an Excel-legal (≤31 chars, no ``[]:*?/\\``), unique sheet name.

    *suffix* (e.g. ``" - Merged"``) is always preserved; *base* is truncated to
    make room, then de-duplicated with ``~2``/``~3`` tags. *used* is mutated with
    the chosen name so callers can thread one set across a workbook.
    """
    cleaned = _ILLEGAL_SHEET.sub(" ", base).strip() or "Sheet"
    suffix = suffix[:31]
    room = max(1, 31 - len(suffix))
    name = (cleaned[:room].strip() + suffix) or "Sheet"
    if name not in used:
        used.add(name)
        return name
    n = 2
    while True:
        tag = f"~{n}"
        trimmed = cleaned[: max(0, room - len(tag))].strip()
        candidate = (trimmed + tag + suffix)[:31]
        if candidate not in used:
            used.add(candidate)
            return candidate
        n += 1


def apply_cell_format(cell: Any, kind: str) -> None:
    """Align + number-format a body cell by its column kind.

    Kinds: ``qty`` (right, integer ``#,##0``) · ``num`` (right, ``#,##0.###`` —
    keeps decimals for chemistry/measurements) · ``inr`` (right, ``₹ #,##0``) ·
    ``credit`` (INR + signed green/red font) · ``date`` (left, ``dd-mm-yyyy``) ·
    ``center`` · anything else ⇒ left text.
    """
    if kind == "qty":
        cell.alignment = ALIGN_RIGHT
        if isinstance(cell.value, (int, float)):
            fmt_quantity(cell)
    elif kind == "num":
        cell.alignment = ALIGN_RIGHT
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.###"
    elif kind == "inr":
        cell.alignment = ALIGN_RIGHT
        if isinstance(cell.value, (int, float)):
            fmt_inr(cell)
    elif kind == "credit":
        cell.alignment = ALIGN_RIGHT
        if isinstance(cell.value, (int, float)):
            fmt_inr(cell)
        apply_credit_balance_font(cell)
    elif kind == "date":
        cell.alignment = ALIGN_LEFT
        if isinstance(cell.value, datetime):
            cell.number_format = "dd-mm-yyyy"
    elif kind == "center":
        cell.alignment = ALIGN_CENTER
    else:  # "text", default
        cell.alignment = ALIGN_LEFT


def write_flat_table(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: Iterable[list[Any]],
    col_formats: dict[int, str] | None = None,
    freeze_cols: int = 0,
    total_row: list[Any] | None = None,
    note: str | None = None,
) -> None:
    """Write one premium flat data table into *ws*.

    *col_formats* maps 1-based column index → kind (see ``apply_cell_format``).
    *total_row* renders an emphasized grand-total row. *note* renders a single
    placeholder row when there are no data rows (e.g. an empty rake breakdown).
    """
    n_cols = len(headers)
    add_title_banner(ws, 1, n_cols, title, subtitle)
    ws.append(headers)
    style_premium_header_row(ws, HEADER_ROW, n_cols)
    _align_header(ws, headers, col_formats)

    body_count = 0
    for i, row in enumerate(rows):
        ws.append(row)
        r = HEADER_ROW + 1 + i
        style_data_row(ws, r, n_cols, is_alt_group=(i % 2 == 1))
        if col_formats:
            for col_idx, kind in col_formats.items():
                apply_cell_format(ws.cell(row=r, column=col_idx), kind)
        body_count += 1

    filter_last = HEADER_ROW + body_count if body_count else HEADER_ROW

    if body_count == 0 and note:
        note_row = HEADER_ROW + 1
        ws.append([note] + [""] * (n_cols - 1))
        ws.cell(row=note_row, column=1).alignment = ALIGN_LEFT
        if n_cols > 1:
            ws.merge_cells(
                start_row=note_row, start_column=1, end_row=note_row, end_column=n_cols
            )

    if total_row is not None:
        ws.append(total_row)
        grand = ws.max_row  # actual appended row (a note row may sit above it)
        style_grand_total_row(ws, grand, n_cols, 1)
        if col_formats:
            for col_idx, kind in col_formats.items():
                apply_cell_format(ws.cell(row=grand, column=col_idx), kind)

    enable_table_filter(ws, HEADER_ROW, n_cols, filter_last)
    freeze_header_and_fixed_cols(ws, HEADER_ROW, freeze_cols)
    ws.row_dimensions[HEADER_ROW].height = 26
    auto_size_columns(ws)
    setup_print_page(ws, title)


def _align_header(
    ws: Worksheet, headers: list[str], col_formats: dict[int, str] | None
) -> None:
    """Right-align numeric header labels, centre status ones (matches body)."""
    if not col_formats:
        return
    for col_idx, kind in col_formats.items():
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        if kind in ("qty", "num", "inr", "credit"):
            cell.alignment = ALIGN_RIGHT
        elif kind == "center":
            cell.alignment = ALIGN_CENTER


def _record_cell(doc: Any, field: str) -> Any:
    """A printable cell value for *field* on a Mongo/pydantic record."""
    value = getattr(doc, field, None)
    if value is None:
        return ""
    # openpyxl rejects tz-aware datetimes (Mongo returns them tz_aware=True).
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def write_records_sheet(
    ws: Worksheet,
    *,
    banner_title: str,
    subtitle: str,
    columns: list[tuple[str, str, str]],
    docs: Iterable[Any],
    freeze_cols: int = 0,
    total_row: list[Any] | None = None,
    note: str | None = None,
) -> None:
    """Write a premium sheet straight from records (Mongo docs / pydantic rows).

    *columns* is ``(header, field, kind)`` per column; ``kind`` drives alignment +
    number format (see ``apply_cell_format``). Rows are read by attribute name.
    """
    headers = [header for header, _, _ in columns]
    col_formats = {i + 1: kind for i, (_, _, kind) in enumerate(columns)}
    rows = ([_record_cell(doc, field) for _, field, _ in columns] for doc in docs)
    write_flat_table(
        ws,
        title=banner_title,
        subtitle=subtitle,
        headers=headers,
        rows=rows,
        col_formats=col_formats,
        freeze_cols=freeze_cols,
        total_row=total_row,
        note=note,
    )


if __name__ == "__main__":  # pragma: no cover — laziest runnable self-check
    # safe_sheet_name: sanitize, truncate (preserving suffix), and de-dup.
    used: set[str] = set()
    assert safe_sheet_name("CWCJ", used, suffix=" - Merged") == "CWCJ - Merged"
    assert safe_sheet_name("CW/CJ:1", used, suffix=" - Merged").startswith("CW CJ 1")
    long = safe_sheet_name("A" * 40, used, suffix=" - Unmerged")
    assert len(long) <= 31 and long.endswith(" - Unmerged")
    a = safe_sheet_name("ROAD", used)
    b = safe_sheet_name("ROAD", used)
    assert a == "ROAD" and b != a and len(b) <= 31

    # write_flat_table: empty body + note + total_row → the grand-total row must be
    # the appended row (row 5), not the note row (row 4). Guards the off-by-one.
    import openpyxl

    ws = openpyxl.Workbook().active
    write_flat_table(
        ws, title="T", subtitle="s", headers=["A", "Total Qty"], rows=[],
        col_formats={2: "qty"}, total_row=["Grand Total", ""], note="No data.",
    )
    assert ws.cell(row=4, column=1).value == "No data."
    assert ws.cell(row=5, column=1).value == "Grand Total"
    assert ws.cell(row=5, column=1).fill.start_color.rgb == "001E293B"  # grand-total fill
    print("excel_premium self-check OK")
