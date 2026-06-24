"""Premium openpyxl styling helpers for the Report JSW/JVML Excel export.

These helpers turn the raw grouped pivot into a CEO-ready workbook: dark themed
header, slate subtotals, alternating group tints, Indian number/currency
formatting, conditional status colours, freeze panes and print-ready layout.
"""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
HEADER_FILL = "0F172A"          # slate-900
HEADER_FONT_COLOUR = "FFFFFF"   # white
SUBTOTAL_FILL = "E2E8F0"        # slate-200
GRAND_TOTAL_FILL = "1E293B"     # slate-800
GRAND_TOTAL_FONT_COLOUR = "FFFFFF"
ALT_GROUP_FILL = "F8FAFC"       # slate-50
DEFAULT_ROW_FILL = "FFFFFF"     # white
BORDER_COLOUR = "D1D5DB"        # gray-300
DARK_BORDER_COLOUR = "9CA3AF"   # gray-400

SEMANTIC_GREEN = "059669"       # emerald-600
SEMANTIC_RED = "DC2626"         # red-600
SEMANTIC_AMBER = "D97706"       # amber-600
MUTED_GREY = "6B7280"           # gray-500

# ---------------------------------------------------------------------------
# Font presets
# ---------------------------------------------------------------------------
FONT_HEADER = Font(bold=True, size=11, color=HEADER_FONT_COLOUR, name="Calibri")
FONT_DATA = Font(size=10, name="Calibri")
FONT_SUBTOTAL = Font(bold=True, size=10, color="1F2937", name="Calibri")
FONT_SUBTOTAL_LABEL = Font(bold=True, size=10, color="1F2937", name="Calibri")
FONT_GRAND_TOTAL = Font(bold=True, size=11, color=GRAND_TOTAL_FONT_COLOUR, name="Calibri")
FONT_TITLE = Font(bold=True, size=16, color=HEADER_FILL, name="Calibri")
FONT_SUBTITLE = Font(size=11, color="475569", name="Calibri")
FONT_STATUS = Font(size=10, bold=True, name="Calibri")

# ---------------------------------------------------------------------------
# Border presets
# ---------------------------------------------------------------------------
THIN_SIDE = Side(style="thin", color=BORDER_COLOUR)
MEDIUM_SIDE = Side(style="medium", color=DARK_BORDER_COLOUR)
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)
SUBTOTAL_TOP = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=MEDIUM_SIDE, bottom=THIN_SIDE
)
GRAND_TOTAL_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=MEDIUM_SIDE,
    bottom=Side(style="double", color=DARK_BORDER_COLOUR),
)

# ---------------------------------------------------------------------------
# Fill presets
# ---------------------------------------------------------------------------
FILL_HEADER = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
FILL_SUBTOTAL = PatternFill(start_color=SUBTOTAL_FILL, end_color=SUBTOTAL_FILL, fill_type="solid")
FILL_GRAND_TOTAL = PatternFill(
    start_color=GRAND_TOTAL_FILL, end_color=GRAND_TOTAL_FILL, fill_type="solid"
)
FILL_ALT_GROUP = PatternFill(
    start_color=ALT_GROUP_FILL, end_color=ALT_GROUP_FILL, fill_type="solid"
)
FILL_DEFAULT = PatternFill(
    start_color=DEFAULT_ROW_FILL, end_color=DEFAULT_ROW_FILL, fill_type="solid"
)

# ---------------------------------------------------------------------------
# Alignment presets
# ---------------------------------------------------------------------------
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_SUBTOTAL_LABEL = Alignment(horizontal="right", vertical="center")


def _apply_to_row(
    ws: Worksheet,
    row_idx: int,
    n_cols: int,
    fill: PatternFill,
    font: Font,
    border: Border,
    alignment: Alignment,
) -> None:
    """Apply the same fill/font/border/alignment to every cell in a row."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = alignment


def style_premium_header_row(ws: Worksheet, row_idx: int, n_cols: int) -> None:
    """Style the header row with a dark theme and white bold text."""
    _apply_to_row(ws, row_idx, n_cols, FILL_HEADER, FONT_HEADER, THIN_BORDER, ALIGN_HEADER)


def style_data_row(
    ws: Worksheet,
    row_idx: int,
    n_cols: int,
    *,
    is_alt_group: bool = False,
) -> None:
    """Style a regular data row with alternating group tint and light borders."""
    fill = FILL_ALT_GROUP if is_alt_group else FILL_DEFAULT
    _apply_to_row(ws, row_idx, n_cols, fill, FONT_DATA, THIN_BORDER, ALIGN_LEFT)


def style_subtotal_row(
    ws: Worksheet,
    row_idx: int,
    n_cols: int,
    label_col_span: int,
) -> None:
    """Style a channel subtotal row: slate fill, bold, medium top border."""
    _apply_to_row(ws, row_idx, n_cols, FILL_SUBTOTAL, FONT_SUBTOTAL, SUBTOTAL_TOP, ALIGN_LEFT)
    # Right-align the subtotal label across the leading columns.
    for col in range(1, label_col_span + 1):
        ws.cell(row=row_idx, column=col).alignment = ALIGN_SUBTOTAL_LABEL


def style_grand_total_row(
    ws: Worksheet,
    row_idx: int,
    n_cols: int,
    label_col_span: int,
) -> None:
    """Style the grand-total row: dark fill, white bold text, double bottom border."""
    _apply_to_row(
        ws, row_idx, n_cols, FILL_GRAND_TOTAL, FONT_GRAND_TOTAL, GRAND_TOTAL_BORDER, ALIGN_LEFT
    )
    for col in range(1, label_col_span + 1):
        ws.cell(row=row_idx, column=col).alignment = ALIGN_SUBTOTAL_LABEL


def fmt_quantity(cell: Any) -> None:
    """Apply Indian-grouped whole-number format to a quantity cell."""
    cell.number_format = '#,##0'


def fmt_inr(cell: Any) -> None:
    """Apply Indian-grouped INR currency format to a money cell."""
    cell.number_format = '₹ #,##0'


def apply_status_font(
    cell: Any,
    status: str,
    *,
    positive: str = SEMANTIC_GREEN,
    negative: str = SEMANTIC_RED,
    warning: str = SEMANTIC_AMBER,
    muted: str = MUTED_GREY,
) -> None:
    """Colour a status cell based on its textual value."""
    text = str(cell.value) if cell.value is not None else ""
    if text in ("Balance Available", "No"):
        cell.font = Font(size=10, bold=True, color=positive, name="Calibri")
    elif text in ("Not Enough Balance", "Balance Negative", "Blocked"):
        cell.font = Font(size=10, bold=True, color=negative, name="Calibri")
    elif text in ("N/A", "NO CREDIT REPORT FOUND", "No Credit balance"):
        cell.font = Font(size=10, italic=True, color=warning, name="Calibri")
    else:
        cell.font = Font(size=10, color=muted, name="Calibri")


def apply_credit_balance_font(cell: Any) -> None:
    """Colour a Credit Balance cell by its signed numeric value or status text."""
    text = str(cell.value) if cell.value is not None else ""

    if text in ("NO CREDIT REPORT FOUND",):
        cell.font = Font(size=10, italic=True, color=SEMANTIC_AMBER, name="Calibri")
        return
    if text in ("No Credit balance",):
        cell.font = Font(size=10, italic=True, color=MUTED_GREY, name="Calibri")
        return

    try:
        value = float(cell.value) if cell.value is not None else None
    except (TypeError, ValueError):
        value = None

    if value is None:
        cell.font = Font(size=10, italic=True, color=MUTED_GREY, name="Calibri")
    elif value >= 0:
        cell.font = Font(size=10, bold=True, color=SEMANTIC_GREEN, name="Calibri")
    else:
        cell.font = Font(size=10, bold=True, color=SEMANTIC_RED, name="Calibri")


def freeze_header_and_fixed_cols(ws: Worksheet, header_row: int, fixed_cols: int) -> None:
    """Freeze the header row and the first *fixed_cols* columns (0 ⇒ rows only)."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=fixed_cols + 1).coordinate


def apply_pivot_view(ws: Worksheet) -> None:
    """Give the sheet a real pivot-table feel: hide the worksheet gridlines so the
    table floats on a clean white canvas, and place each collapsible group's +/-
    summary on the subtotal row that sits *below* its detail rows."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_properties.outlinePr.applyStyles = True


def setup_print_page(ws: Worksheet, title: str) -> None:
    """Configure the worksheet for professional printing: landscape, fit-to-width,
    A4 paper, and repeat the top row on every printed page."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.title = title


def add_title_banner(
    ws: Worksheet,
    row_idx: int,
    n_cols: int,
    title: str,
    subtitle: str,
) -> None:
    """Add a merged title banner above the table."""
    start_cell = ws.cell(row=row_idx, column=1, value=title)
    start_cell.font = FONT_TITLE
    start_cell.alignment = Alignment(horizontal="left", vertical="center")

    subtitle_cell = ws.cell(row=row_idx + 1, column=1, value=subtitle)
    subtitle_cell.font = FONT_SUBTITLE
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")

    if n_cols > 1:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)
        ws.merge_cells(
            start_row=row_idx + 1, start_column=1, end_row=row_idx + 1, end_column=n_cols
        )

    ws.row_dimensions[row_idx].height = 28
    ws.row_dimensions[row_idx + 1].height = 20


def auto_size_columns(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
    """Resize every column to the widest populated cell, clamped to bounds."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells
        )
        width = max(min_width, min(max_width, max_len + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def enable_table_filter(ws: Worksheet, header_row: int, n_cols: int, last_row: int) -> None:
    """Enable auto-filter only over the data table (header + body rows)."""
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{last_row}"


def add_premium_metadata_sheet(
    wb: Workbook,
    title: str,
    items: dict[str, str],
) -> None:
    """Append a styled Metadata sheet with a title and bordered key/value rows."""
    ws = wb.create_sheet(title="Metadata")

    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Key"
    ws["B3"] = "Value"
    style_premium_header_row(ws, 3, 2)

    for row_idx, (key, value) in enumerate(items.items(), start=4):
        key_cell = ws.cell(row=row_idx, column=1, value=key)
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        for cell in (key_cell, val_cell):
            cell.font = FONT_DATA
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_LEFT
        key_cell.font = Font(size=10, bold=True, name="Calibri")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55
