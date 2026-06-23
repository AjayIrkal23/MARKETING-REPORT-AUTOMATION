"""Excel import and template utilities for the customer_code domain.

``openpyxl`` is imported **lazily inside each function** so ``import app.main``
works without the package installed.  Real-file quirks handled: ``code`` as int,
``MOB No.`` as int or free-text, ``CAM `` trailing space, numeric cells coerced
to strings, and strict template enforcement via a hidden fingerprint sheet.
"""

from __future__ import annotations

import re

from ...schemas.customer_code import CustomerCodeImportError


# ---------------------------------------------------------------------------
# Template fingerprint
# ---------------------------------------------------------------------------

#: Hidden worksheet name used to mark an official portal-generated workbook.
_TEMPLATE_MARKER_SHEET: str = "_JSW_MRA_TEMPLATE_"

#: Cell values stored in the hidden fingerprint sheet.
_TEMPLATE_MARKER_A1: str = "CUSTOMER_CODES_TEMPLATE"
_TEMPLATE_MARKER_A2: str = "v1"


# ---------------------------------------------------------------------------
# Header normalisation
# ---------------------------------------------------------------------------

def normalize_header(h: object) -> str:
    """Collapse whitespace runs to a single space, strip, and lower-case.

    Ensures ``"CAM "`` → ``"cam"``, ``"MOB No."`` → ``"mob no."`` etc.
    """
    return re.sub(r"\s+", " ", str(h)).strip().lower()


# ---------------------------------------------------------------------------
# Column → field map
# ---------------------------------------------------------------------------

#: Maps *normalised* Excel header strings to CustomerCode model field names.
#: Three ``mob`` aliases handle period-dropped exports and bare ``"mob"``.
_HEADER_MAP: dict[str, str] = {
    "segment":          "segment",
    "code":             "code",
    "customer":         "customer",
    "destination":      "destination",
    "cam":              "cam",            # 'CAM ' → 'cam' via normalize_header
    "mob no.":          "mob",            # 'MOB No.' → 'mob no.'
    "mob no":           "mob",            # alias: period dropped by some exports
    "mob":              "mob",            # bare alias
    "head":             "head",
    "route":            "route",
    "ship to":          "ship_to",
    "ship to customer": "ship_to_customer",
    "ship to city":     "ship_to_city",
    "rake":             "rake",
    "transport mode":   "transport_mode",
}

#: Required model fields — a non-empty row missing any of these is an error.
REQUIRED_FIELDS: tuple[str, ...] = ("segment", "code", "customer", "destination")

#: Maximum data rows accepted per import sheet (DoS guard).
MAX_IMPORT_ROWS: int = 50_000


# ---------------------------------------------------------------------------
# Header resolution
# ---------------------------------------------------------------------------

def _header_to_field(header_row: tuple[object, ...]) -> dict[int, str]:
    """Build a column-index → field-name map from the header row.

    Every non-empty header must exist in ``_HEADER_MAP``; callers enforce
    template conformance before this function is reached.
    """
    col_field_map: dict[int, str] = {}
    for col_idx, h in enumerate(header_row):
        if h is None:
            continue
        norm = normalize_header(h)
        field = _HEADER_MAP.get(norm)
        if field:
            col_field_map[col_idx] = field
    return col_field_map


def _strip_trailing_nones(header_row: tuple[object, ...]) -> list[object]:
    """Return the header row with trailing ``None`` cells removed."""
    headers = list(header_row)
    while headers and headers[-1] is None:
        headers.pop()
    return headers


# ---------------------------------------------------------------------------
# Cell coercion
# ---------------------------------------------------------------------------

def cell_to_str(v: object) -> str | None:
    """Coerce a raw openpyxl cell value to a clean string or ``None``.

    - ``None`` → ``None``
    - ``float`` with no fractional part (e.g. ``40020365.0``) → ``"40020365"``
    - ``float`` with fractional part → ``str(v)``
    - ``int`` / ``str`` / other → ``str(v).strip()``; empty string → ``None``

    The ``float`` branch handles SAP codes stored as numeric xlsx cells
    (openpyxl reads them as ``float`` in ``data_only`` mode).
    """
    if v is None:
        return None
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    s = str(v).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Workbook parser
# ---------------------------------------------------------------------------

def _verify_template_fingerprint(wb) -> bool:
    """Return True if the workbook contains the official portal fingerprint."""
    if _TEMPLATE_MARKER_SHEET not in wb.sheetnames:
        return False
    ws = wb[_TEMPLATE_MARKER_SHEET]
    return (
        ws["A1"].value == _TEMPLATE_MARKER_A1
        and ws["A2"].value == _TEMPLATE_MARKER_A2
    )


def _validate_headers(
    header_row: tuple[object, ...],
) -> tuple[dict[int, str] | None, CustomerCodeImportError | None]:
    """Check header conformance and return ``(col_field_map, error)``.

    The visible headers must be a subsequence of ``TEMPLATE_HEADERS`` after
    normalisation: order must be preserved, no unknown/extra headers, and no
    duplicate headers.  Missing columns are allowed; required columns that are
    absent will have their values default to ``"unknown"`` at the row level.
    """
    actual_headers = _strip_trailing_nones(header_row)
    expected_norm = [normalize_header(h) for h in TEMPLATE_HEADERS]
    actual_norm = [normalize_header(h) for h in actual_headers]

    # Detect unknown / extra headers.
    unknown = [h for h in actual_norm if h not in expected_norm]
    if unknown:
        return None, CustomerCodeImportError(
            row=0,
            message=(
                "Uploaded file contains unexpected column(s): "
                f"{', '.join(unknown)}. "
                "Download the official customer codes template and use it without modification."
            ),
        )

    # Detect duplicate headers.
    seen = set()
    duplicates = {h for h in actual_norm if h in seen or seen.add(h)}  # type: ignore[func-returns-value]
    if duplicates:
        return None, CustomerCodeImportError(
            row=0,
            message=(
                "Uploaded file contains duplicate column(s): "
                f"{', '.join(sorted(duplicates))}. "
                "Download the official customer codes template and use it without modification."
            ),
        )

    # Require the actual headers to appear in the same order as the template.
    expected_iter = iter(expected_norm)
    for h in actual_norm:
        for expected in expected_iter:
            if h == expected:
                break
        else:
            # `h` was not found in the remaining expected headers → wrong order.
            return None, CustomerCodeImportError(
                row=0,
                message=(
                    "Uploaded file columns are out of order. "
                    "Download the official customer codes template and use it without modification."
                ),
            )

    return _header_to_field(header_row), None


def parse_workbook(
    file_bytes: bytes,
) -> tuple[list[dict[str, str | None]], list[CustomerCodeImportError]]:
    """Parse binary ``.xlsx`` bytes and return ``(valid_rows, errors)``.

    Each element of ``valid_rows`` is a ``dict`` of model field → cleaned
    string (or ``None`` for optional fields).  Fully-empty rows are silently
    skipped (counted as *skipped*, not as errors in the import summary).
    Missing columns are allowed; absent required columns result in ``"unknown"``
    for every row, and absent optional columns result in ``None``.

    Early-return paths:
    (1) missing/invalid fingerprint → row-0 error;
    (2) unknown, duplicate, or out-of-order headers → row-0 error;
    (3) no rows → ``([], [])``;
    (4) exceeds ``MAX_IMPORT_ROWS`` → row-0 error.

    Excel row numbers: header = 1, first data row = 2.
    Rows buffered before ``wb.close()`` — openpyxl read-only mode
    releases file handles on close.
    """
    import io

    import openpyxl  # lazy — app.main works without openpyxl installed

    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )

    if not _verify_template_fingerprint(wb):
        wb.close()
        return [], [
            CustomerCodeImportError(
                row=0,
                message=(
                    "Uploaded file is not the official customer codes template. "
                    "Download the template from this portal and use it without modification."
                ),
            )
        ]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))  # buffer before close
    wb.close()  # required: releases file handles held by read_only mode

    if not rows:
        return [], []

    # Validate headers and build column-index → field-name map.
    col_field_map, header_error = _validate_headers(rows[0])
    if header_error is not None or col_field_map is None:
        return [], [header_error] if header_error is not None else []

    # DoS guard — reject sheets that exceed MAX_IMPORT_ROWS.
    data_rows = rows[1:]
    if len(data_rows) > MAX_IMPORT_ROWS:
        return [], [
            CustomerCodeImportError(
                row=0,
                message=(
                    f"Sheet exceeds {MAX_IMPORT_ROWS:,} row limit "
                    f"({len(data_rows):,} rows found). Split the file."
                ),
            )
        ]

    valid_rows: list[dict[str, str | None]] = []
    errors: list[CustomerCodeImportError] = []

    all_fields = set(_HEADER_MAP.values())

    for row_idx, raw_row in enumerate(data_rows, start=2):
        # Seed every known field with None, then override with actual values.
        # Missing columns therefore produce explicit None values rather than
        # absent keys; required fields are filled with "unknown" below.
        parsed: dict[str, str | None] = {field: None for field in all_fields}
        parsed.update({
            field: cell_to_str(raw_row[col_i] if col_i < len(raw_row) else None)
            for col_i, field in col_field_map.items()
        })

        # Branch A: all mapped fields are None → fully-empty row, skip silently.
        # These contribute to the *skipped* count, not the error count.
        if all(v is None for v in parsed.values()):
            continue

        # Branch B: non-empty row missing a required field → fill with "unknown"
        # rather than rejecting the row. The user explicitly wants every row kept.
        for field in REQUIRED_FIELDS:
            if not parsed.get(field):
                parsed[field] = "unknown"

        valid_rows.append(parsed)

    return valid_rows, errors


# ---------------------------------------------------------------------------
# Template builder
# ---------------------------------------------------------------------------

#: Canonical column order for the downloadable import template.
#: Written without the trailing space on "CAM"; normalize_header handles
#: the equivalence on re-import.
TEMPLATE_HEADERS: list[str] = [
    "Segment", "code", "Customer", "Destination",
    "CAM", "MOB No.", "Head", "ROUTE",
    "SHIP TO", "SHIP TO CUSTOMER",
    "SHIP TO CITY", "RAKE", "TRANSPORT MODE",
]

_TEMPLATE_EXAMPLE_ROW: list[str] = [
    "Retail", "40020365", "Example Steel Traders", "Mumbai",
    "Ravi Kumar", "9876543210", "Sales Head Name", "KAT036",
    "40047421", "Example Ship-To Pvt Ltd",
    "Mumbai", "RAKE-01", "RAKE",
]


def add_template_fingerprint(wb) -> None:
    """Attach the official-portal fingerprint sheet to ``wb``.

    The sheet is hidden and contains a stable marker.  It is preserved when
    the user edits the workbook in Excel/LibreOffice, but lost when values are
    copied into a new blank workbook — this enforces template-only imports.
    """
    ws = wb.create_sheet(title=_TEMPLATE_MARKER_SHEET)
    ws["A1"] = _TEMPLATE_MARKER_A1
    ws["A2"] = _TEMPLATE_MARKER_A2
    ws.sheet_state = "hidden"


def build_template_workbook() -> bytes:
    """Build a minimal import template and return raw ``.xlsx`` bytes.

    The workbook contains one visible sheet (``"Customer Codes"``) with the
    canonical header row and one example data row illustrating expected value
    formats, plus a hidden fingerprint sheet proving it was generated by the
    portal.  Headers are bold with a light-blue fill for readability; columns
    are auto-sized to the widest of header or example value.

    Returns raw bytes suitable for ``StreamingResponse`` with content-type
    ``application/vnd.openxmlformats-officedocument.spreadsheetml.sheet``.

    ``openpyxl`` is imported lazily so app startup does not require it.
    """
    import io

    import openpyxl  # lazy — app.main works without openpyxl installed
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Codes"

    add_template_fingerprint(wb)

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws.append(_TEMPLATE_EXAMPLE_ROW)

    # Auto-size columns to the widest of header or example value.
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
