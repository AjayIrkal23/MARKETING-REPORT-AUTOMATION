"""Unit test for Report JSW/JVML RAKE-pivot Excel export.

Mocks the report generator so no MongoDB connection is required.
"""

from __future__ import annotations

import asyncio
import io
from unittest.mock import AsyncMock, patch

from openpyxl import load_workbook

from app.schemas.report import (
    CombinedExportQuery,
    ReportPivotRow,
    ReportQuery,
    ReportResponse,
)
from app.services.report.export_combined import export_combined


def _fake_report() -> ReportResponse:
    row = ReportPivotRow(
        so_sales_org="1001",
        distr_chnl="OEM",
        sold_to_party="ACME Steels Ltd",
        sales_office="Mumbai",
        party_code="40122581",
        ship_to_party="ACME Ship-To",
        transport_mode="ROAD/RAKE",
        destination="Jaipur",
        route="KAT138",
        rake_quantities={"KKU": 55.01, "ROAD": 0.0},
        total=55.01,
        nco_yes_do=0.0,
        nco_yes_do_count=0,
        blocked=False,
        credit_balance=500000.0,
        required_credit=55.01,
        credit_sufficient=True,
        credit_status="",
        credit_note="Balance Available",
    )
    return ReportResponse(
        date="23-06-2026",
        report_type="jsw",
        region_id=None,
        region_name="All Regions",
        days_filter="include",
        ccas=["VJ0H", "1000"],
        has_stock=True,
        has_credit_report=True,
        coil_price_per_qty=1.0,
        rake_columns=["KKU", "ROAD"],
        rows=[row],
        grand_total=55.01,
        grand_nco_yes_do=0.0,
        grand_required_credit=55.01,
    )


def test_export_builds_xlsx_bytes() -> None:
    """export_report should return non-empty .xlsx bytes."""
    with patch(
        "app.services.report.export_combined.generate_report",
        new=AsyncMock(return_value=_fake_report()),
    ):
        query = CombinedExportQuery(date="23-06-2026", report_type="jsw", sheets="pivot")
        data = asyncio.run(export_combined(query))

    assert isinstance(data, bytes)
    assert len(data) > 0
    assert data.startswith(b"PK")  # .xlsx is a zip archive


def _mk_row(*, channel: str, party: str, total: float, org: str = "1001") -> ReportPivotRow:
    return ReportPivotRow(
        so_sales_org=org,
        distr_chnl=channel,
        sold_to_party="ACME Steels Ltd",
        sales_office="Mumbai",
        party_code=party,
        ship_to_party="ACME Ship-To",
        transport_mode="ROAD",
        destination="Jaipur",
        route="KAT138",
        rake_quantities={"KKU": total, "ROAD": 0.0},
        total=total,
        nco_yes_do=0.0,
        nco_yes_do_count=0,
        blocked=False,
        credit_balance=500000.0,
        required_credit=total,
        credit_sufficient=True,
        credit_status="",
        credit_note="Balance Available",
    )


def test_export_groups_with_subtotals_and_blanked_parents() -> None:
    """Two OEM rows + one TRADE row → per-channel subtotals, blanked repeated parent, grand total."""
    rows = [
        _mk_row(channel="OEM", party="40000001", total=10.0),
        _mk_row(channel="OEM", party="40000002", total=20.0),
        _mk_row(channel="TRADE", party="40000003", total=5.0),
    ]
    report = _fake_report()
    report.rows = rows
    report.grand_total = 35.0
    report.grand_required_credit = 35.0

    with patch(
        "app.services.report.export_combined.generate_report",
        new=AsyncMock(return_value=report),
    ):
        data = asyncio.run(export_combined(CombinedExportQuery(date="23-06-2026", report_type="jsw", sheets="pivot")))

    ws = load_workbook(io.BytesIO(data))["BRANCH WISE PIVOT REPORT"]
    grid = [[c.value for c in r] for r in ws.iter_rows()]
    header = grid[2]  # title banner at rows 1-2, table header at row 3
    total_col = header.index("Total")
    distr_col = header.index("Distr. Channel")
    first_col = [r[0] for r in grid]

    # Per-channel subtotal rows exist with the right summed Total.
    oem_sub = next(r for r in grid if r[0] == "OEM Total")
    trade_sub = next(r for r in grid if r[0] == "TRADE Total")
    assert oem_sub[total_col] == 30.0
    assert trade_sub[total_col] == 5.0

    # Grand total row.
    grand = next(r for r in grid if r[0] == "Grand Total")
    assert grand[total_col] == 35.0

    # The 2nd OEM data row repeats the channel → its Distr.Channel cell is blanked.
    oem_data = [r for r in grid if r[total_col] in (10.0, 20.0)]
    assert oem_data[0][distr_col] == "OEM"          # first row shows the channel
    assert oem_data[1][distr_col] in (None, "")     # second row blanked (empty cell)
    # Subtotals appear before the grand total.
    assert first_col.index("OEM Total") < first_col.index("Grand Total")


def _export_headers(columns: str | None) -> list[str]:
    report = _fake_report()
    report.rows = [_mk_row(channel="OEM", party="40000001", total=10.0)]
    query = CombinedExportQuery(date="23-06-2026", report_type="jsw", columns=columns, sheets="pivot")
    with patch(
        "app.services.report.export_combined.generate_report",
        new=AsyncMock(return_value=report),
    ):
        data = asyncio.run(export_combined(query))
    ws = load_workbook(io.BytesIO(data))["BRANCH WISE PIVOT REPORT"]
    return [c.value for c in ws[3]]


def test_export_respects_columns_filter() -> None:
    """The export honours the optional-column filter, including Total as optional."""
    # Only the 5 fixed left cols are always shown; the RAKE block is gated by "rake".
    always_on = {"Distr. Channel", "BRANCH", "Sold To Party", "Party Code", "Ship To Party"}

    # Filter to {total, route}: those two optional cols show; the rest (incl. RAKE block) drop.
    headers = _export_headers("total,route")
    assert always_on <= set(headers)
    assert "ROUTE" in headers and "Total" in headers
    for dropped in ("Transport Mode", "Destination", "Yes+DO", "Blocked", "Credit Balance", "Required Credit", "Credit Note", "KKU", "ROAD"):
        assert dropped not in headers

    # The whole dynamic RAKE block is shown/hidden by the single "rake" key.
    assert {"KKU", "ROAD"} <= set(_export_headers("rake"))

    # Empty filter → only the always-on columns (every optional col + RAKE block hidden).
    bare = _export_headers("")
    assert set(bare) == always_on

    # None (param absent) → every column, Total present.
    full = _export_headers(None)
    assert "Total" in full and "Transport Mode" in full and "Credit Note" in full

    # Hiding Total drops it (it is genuinely optional now).
    no_total = _export_headers("yes_do")
    assert "Total" not in no_total and "Yes+DO" in no_total


def test_export_branch_grouped_after_distr_channel() -> None:
    """BRANCH is the grouped pivot column immediately after Distr. Channel (above Sold To Party)."""
    headers = _export_headers(None)
    assert headers.index("BRANCH") == headers.index("Distr. Channel") + 1
    assert headers.index("BRANCH") < headers.index("Sold To Party")


def _styled_ws():
    """Return the exported worksheet with two channel groups and varied credit states."""
    rows = [
        ReportPivotRow(
            so_sales_org="1001",
            distr_chnl="OEM",
            sold_to_party="Alpha Steels",
            sales_office="Mumbai",
            party_code="40000001",
            ship_to_party="Alpha Ship-To",
            transport_mode="ROAD",
            destination="Pune",
            route="KAT001",
            rake_quantities={"KKU": 10.0},
            total=10.0,
            nco_yes_do=2.0,
            nco_yes_do_count=1,
            blocked=False,
            credit_balance=500000.0,
            required_credit=8.0,
            credit_sufficient=True,
            credit_status="",
            credit_note="Balance Available",
        ),
        ReportPivotRow(
            so_sales_org="1001",
            distr_chnl="OEM",
            sold_to_party="Beta Steels",
            sales_office="Mumbai",
            party_code="40000002",
            ship_to_party="Beta Ship-To",
            transport_mode="RAKE",
            destination="Pune",
            route="KAT002",
            rake_quantities={"KKU": 20.0},
            total=20.0,
            nco_yes_do=0.0,
            nco_yes_do_count=0,
            blocked=True,
            credit_balance=-10000.0,
            required_credit=20.0,
            credit_sufficient=False,
            credit_status="",
            credit_note="Balance Negative",
        ),
        ReportPivotRow(
            so_sales_org="1001",
            distr_chnl="TRADE",
            sold_to_party="Gamma Steels",
            sales_office="Delhi",
            party_code="40000003",
            ship_to_party="Gamma Ship-To",
            transport_mode="ROAD",
            destination="Jaipur",
            route="KAT003",
            rake_quantities={"ROAD": 5.0},
            total=5.0,
            nco_yes_do=0.0,
            nco_yes_do_count=0,
            blocked=False,
            credit_balance=1000.0,
            required_credit=50.0,
            credit_sufficient=False,
            credit_status="",
            credit_note="Not Enough Balance",
        ),
    ]
    report = _fake_report()
    report.rows = rows
    report.rake_columns = ["KKU", "ROAD"]
    report.grand_total = 35.0
    report.grand_nco_yes_do = 2.0
    report.grand_required_credit = 78.0

    with patch(
        "app.services.report.export_combined.generate_report",
        new=AsyncMock(return_value=report),
    ):
        data = asyncio.run(export_combined(CombinedExportQuery(date="23-06-2026", report_type="jsw", sheets="pivot")))

    return load_workbook(io.BytesIO(data))["BRANCH WISE PIVOT REPORT"]


def test_export_has_premium_title_banner() -> None:
    """The report sheet should open with a merged title and subtitle banner."""
    ws = _styled_ws()
    assert ws["A1"].value == "JSW Coil Stock Report"
    assert "Report Date: 23-06-2026" in ws["A2"].value
    assert ws["A1"].font.bold is True


def test_export_header_is_dark_themed() -> None:
    """The table header should have a dark fill and white bold text."""
    ws = _styled_ws()
    header_row = 3
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        assert cell.fill.start_color.rgb == "000F172A"
        assert cell.font.color.rgb == "00FFFFFF"
        assert cell.font.bold is True


def test_export_subtotal_and_grand_total_styling() -> None:
    """Subtotal rows should be slate/bold; the grand total should be dark/white."""
    ws = _styled_ws()
    grid = [[c.value for c in r] for r in ws.iter_rows()]

    oem_sub_idx = next(i for i, r in enumerate(grid, start=1) if r[0] == "OEM Total")
    grand_idx = next(i for i, r in enumerate(grid, start=1) if r[0] == "Grand Total")

    assert ws.cell(row=oem_sub_idx, column=1).fill.start_color.rgb == "00E2E8F0"
    assert ws.cell(row=oem_sub_idx, column=1).font.bold is True

    assert ws.cell(row=grand_idx, column=1).fill.start_color.rgb == "001E293B"
    assert ws.cell(row=grand_idx, column=1).font.color.rgb == "00FFFFFF"
    assert ws.cell(row=grand_idx, column=1).font.bold is True


def test_export_numeric_columns_formatted() -> None:
    """RAKE, Total and Yes+DO columns should use Indian grouped whole-number format."""
    ws = _styled_ws()
    header = [c.value for c in ws[3]]
    kku_col = header.index("KKU") + 1
    total_col = header.index("Total") + 1

    # Data rows use the quantity format.
    assert ws.cell(row=4, column=kku_col).number_format == '#,##0'
    assert ws.cell(row=4, column=total_col).number_format == '#,##0'


def test_export_inr_columns_formatted() -> None:
    """Credit Balance and Required Credit should use the ₹ Indian-grouped format."""
    ws = _styled_ws()
    header = [c.value for c in ws[3]]
    req_col = header.index("Required Credit") + 1
    assert ws.cell(row=4, column=req_col).number_format == '₹ #,##0'


def test_export_status_columns_have_conditional_colours() -> None:
    """Blocked and Credit Note cells should be coloured by their semantic state."""
    ws = _styled_ws()
    header = [c.value for c in ws[3]]
    blocked_col = header.index("Blocked") + 1
    note_col = header.index("Credit Note") + 1

    # First row: No / Balance Available → green.
    assert ws.cell(row=4, column=blocked_col).font.color.rgb == "00059669"
    assert ws.cell(row=4, column=note_col).font.color.rgb == "00059669"

    # Second row: Blocked / Balance Negative → red.
    assert ws.cell(row=5, column=blocked_col).font.color.rgb == "00DC2626"
    assert ws.cell(row=5, column=note_col).font.color.rgb == "00DC2626"


def test_export_credit_balance_conditional_colour() -> None:
    """Credit Balance cells should be green for positive and red for negative values."""
    ws = _styled_ws()
    header = [c.value for c in ws[3]]
    bal_col = header.index("Credit Balance") + 1

    assert ws.cell(row=4, column=bal_col).font.color.rgb == "00059669"  # 500000
    assert ws.cell(row=5, column=bal_col).font.color.rgb == "00DC2626"  # -10000


def test_export_freeze_panes_and_print_setup() -> None:
    """Header rows freeze; NO columns are frozen; sheet is print-ready."""
    ws = _styled_ws()
    # Freeze pane is set at A4 (below header row 3) — no fixed/frozen columns.
    assert ws.freeze_panes == "A4"
    assert ws.page_setup.orientation == "landscape"
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.print_title_rows == "$1:$1"


def test_export_pivot_view() -> None:
    """Pivot-table feel: gridlines hidden, collapsible groups with summary below,
    and detail rows sit one outline level under their subtotal."""
    ws = _styled_ws()
    assert ws.sheet_view.showGridLines is False
    assert ws.sheet_properties.outlinePr.summaryBelow is True
    # At least one detail row is grouped (outline level 1); subtotal/grand stay 0.
    levels = {ws.row_dimensions[r].outline_level for r in range(4, ws.max_row + 1)}
    assert 1 in levels


def test_merge_reports_groups_by_so_org() -> None:
    """`both` merges jsw + jvml: rows sorted by SO Sales Org, union RAKE cols, summed grands."""
    from app.services.report.generate import _merge_reports

    jsw = _fake_report()  # org 1001, ccas [VJ0H, 1000]
    jsw.rows = [_mk_row(channel="OEM", party="40000001", total=10.0, org="1001")]
    jsw.rake_columns = ["KKU"]
    jsw.grand_total = 10.0

    jvml = _fake_report()
    jvml.report_type = "jvml"
    jvml.ccas = ["JV0H"]
    jvml_row = _mk_row(channel="OEM", party="40000099", total=7.0, org="1060")
    jvml_row.rake_quantities = {"CWCJ": 7.0}
    jvml.rows = [jvml_row]
    jvml.rake_columns = ["CWCJ"]
    jvml.grand_total = 7.0

    merged = _merge_reports(ReportQuery(date="23-06-2026", report_type="both"), [jsw, jvml])

    assert merged.report_type == "both"
    assert merged.rake_columns == ["CWCJ", "KKU"]  # sorted union of both reports
    assert [r.so_sales_org for r in merged.rows] == ["1001", "1060"]  # grouped by SO org
    assert merged.ccas == ["VJ0H", "1000", "JV0H"]
    assert merged.grand_total == 17.0
    assert merged.has_credit_report is True


def test_export_both_leads_with_so_org_column_and_subtotals() -> None:
    """`both` export: SO Sales Org is the first column with a per-SO-org subtotal row."""
    report = _fake_report()
    report.report_type = "both"
    report.rows = [
        _mk_row(channel="OEM", party="40000001", total=10.0, org="1001"),
        _mk_row(channel="OEM", party="40000099", total=7.0, org="1060"),
    ]
    report.rake_columns = ["KKU"]
    report.grand_total = 17.0
    report.grand_required_credit = 17.0

    with patch(
        "app.services.report.export_combined.generate_report",
        new=AsyncMock(return_value=report),
    ):
        data = asyncio.run(export_combined(CombinedExportQuery(date="23-06-2026", report_type="both", sheets="pivot")))

    ws = load_workbook(io.BytesIO(data))["BRANCH WISE PIVOT REPORT"]
    assert [c.value for c in ws[3]][0] == "SO Sales Org"  # header row 3, first col
    first_col = [ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)]
    assert "1001 Total" in first_col
    assert "1060 Total" in first_col
