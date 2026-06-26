"""Unit test for JVML Stock Excel export.

Mocks the Beanie query layer so no MongoDB connection is required.
"""

from __future__ import annotations

import asyncio
import io
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

from app.schemas.jvml_stock import JvmlStockListQuery
from app.services.jvml_stock.export import export_jvml_stock


def _make_doc(**kwargs: object) -> SimpleNamespace:
    """Build a fake JvmlStock document for export testing."""
    defaults = {
        "report_date": "23-06-2026",
        "party_code": "000040122581",
        "party_code_normalized": "40122581",
        "customer_name": "ACME Steels",
        "sold_to_party": "ACME Steels Ltd",
        "material": "HR Coil",
        "jsw_grade": "SAE1006",
        "sales_office": "MH01",
        "distr_chnl": "OEM",
        "unrestr_qty": 10.0,
        "stock_quantity": 12.0,
    }
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


def test_export_builds_xlsx_bytes() -> None:
    """export_jvml_stock should return non-empty .xlsx bytes."""
    fake_doc = _make_doc()

    mock_find = MagicMock()
    mock_find.sort = MagicMock(return_value=mock_find)
    mock_find.to_list = AsyncMock(return_value=[fake_doc])

    with patch(
        "app.services.jvml_stock.export.JvmlStock.find",
        return_value=mock_find,
    ):
        with patch(
            "app.services.shared.stock_export.CustomerCode.find",
            return_value=MagicMock(to_list=AsyncMock(return_value=[]))(),
        ):
            query = JvmlStockListQuery(date="23-06-2026")
            data = asyncio.run(export_jvml_stock(query))

    assert isinstance(data, bytes)
    assert len(data) > 0
    assert data.startswith(b"PK")  # .xlsx is a zip archive


def test_export_includes_all_client_columns_and_uses_normalized_party_code() -> None:
    """The export workbook should contain every client-facing source column,
    the resolved customer name and report date, the normalized party code, and
    no system-generated columns.
    """
    fake_doc = _make_doc()

    mock_find = MagicMock()
    mock_find.sort = MagicMock(return_value=mock_find)
    mock_find.to_list = AsyncMock(return_value=[fake_doc])

    with patch(
        "app.services.jvml_stock.export.JvmlStock.find",
        return_value=mock_find,
    ):
        with patch(
            "app.services.shared.stock_export.CustomerCode.find",
            return_value=MagicMock(to_list=AsyncMock(return_value=[]))(),
        ):
            query = JvmlStockListQuery(date="23-06-2026")
            data = asyncio.run(export_jvml_stock(query))

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["JVML Stock"]
    headers = [cell.value for cell in ws[3]]  # rows 1-2 = banner, header on row 3

    assert "Party Code (Normalized)" in headers
    assert "Customer Name" in headers
    assert "Report Date" in headers
    assert "SO Sales Org" in headers
    assert "Route Desc" in headers

    # Raw zero-padded party_code is not exported; system columns are excluded.
    assert "party_code" not in headers
    assert "id" not in headers
    assert "row_hash" not in headers
    assert "customer_code_id" not in headers
    assert "source_file" not in headers
    assert "created_at" not in headers
    assert "updated_at" not in headers

    # The normalized value should be written under the normalized header.
    normalized_idx = headers.index("Party Code (Normalized)") + 1
    assert ws.cell(row=4, column=normalized_idx).value == "40122581"  # first data row
