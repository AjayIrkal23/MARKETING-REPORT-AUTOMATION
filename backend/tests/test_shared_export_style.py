"""Pure-unit tests for the shared Excel export styling helpers."""

from __future__ import annotations

import io

import openpyxl

from app.utils.shared.export_style import (
    add_metadata_sheet,
    auto_size_columns,
    enable_filters,
    style_header_row,
)


def _load_workbook(data: bytes):
    """Reload a workbook from in-memory bytes to verify saved state."""
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


def test_style_header_row_applies_bold_fill_and_borders() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Segment", "Code", "Customer"])

    style_header_row(ws[1])

    for cell in ws[1]:
        assert cell.font.bold is True
        assert cell.fill.fill_type == "solid"
        assert cell.fill.start_color.rgb.endswith("D9E1F2")
        assert cell.border.left.border_style == "thin"
        assert cell.border.right.border_style == "thin"
        assert cell.border.top.border_style == "thin"
        assert cell.border.bottom.border_style == "thin"


def test_enable_filters_sets_auto_filter_reference() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append(["1", "2"])

    enable_filters(ws)

    assert ws.auto_filter.ref == ws.dimensions


def test_auto_size_columns_clamps_widths() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Short", "A" * 200])

    auto_size_columns(ws, min_width=10, max_width=60)

    # Column A should be wider than min but modest.
    assert ws.column_dimensions["A"].width >= 10
    # Column B must respect the max_width cap.
    assert ws.column_dimensions["B"].width <= 60


def test_add_metadata_sheet_creates_title_and_key_values() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Data"])

    add_metadata_sheet(
        wb,
        title="Test Report",
        items={"Exported At": "2026-01-01T00:00:00Z", "Filter Summary": "None"},
    )

    assert "Metadata" in wb.sheetnames
    meta = wb["Metadata"]
    assert meta["A1"].value == "Test Report"
    assert meta["A3"].value == "Key"
    assert meta["B3"].value == "Value"
    assert meta["A4"].value == "Exported At"
    assert meta["B4"].value == "2026-01-01T00:00:00Z"


def test_helpers_round_trip_through_save_and_load() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Segment", "Code", "Customer"])
    style_header_row(ws[1])
    ws.append(["Retail", "40020365", "Example"])
    enable_filters(ws)
    auto_size_columns(ws)
    add_metadata_sheet(wb, title="Round Trip", items={"Rows": "1"})

    buf = io.BytesIO()
    wb.save(buf)
    loaded = _load_workbook(buf.getvalue())

    data_ws = loaded.active
    header = data_ws[1]
    assert header[0].font.bold is True
    assert header[0].fill.fill_type == "solid"
    assert data_ws.auto_filter.ref is not None
    assert "Metadata" in loaded.sheetnames
