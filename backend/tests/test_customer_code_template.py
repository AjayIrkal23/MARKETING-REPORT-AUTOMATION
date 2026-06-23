"""Pure-unit tests for the customer-code Excel template and import parser.

No network, no MongoDB, no async.  Verifies the official template fingerprint,
header conformance, and field mapping.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from app.schemas.customer_code import CustomerCodeImportError
from app.utils.customer_code.excel import (
    TEMPLATE_HEADERS,
    build_template_workbook,
    parse_workbook,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_workbook(
    headers: list[str],
    data_rows: list[list[str | None]] | None = None,
    *,
    include_fingerprint: bool = True,
) -> bytes:
    """Create an in-memory .xlsx workbook for parser tests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Codes"

    if include_fingerprint:
        fp = wb.create_sheet(title="_JSW_MRA_TEMPLATE_")
        fp["A1"] = "CUSTOMER_CODES_TEMPLATE"
        fp["A2"] = "v1"
        fp.sheet_state = "hidden"

    ws.append(headers)
    for row in data_rows or []:
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Template builder
# ---------------------------------------------------------------------------

def test_template_contains_fingerprint_sheet() -> None:
    data = build_template_workbook()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    assert "_JSW_MRA_TEMPLATE_" in wb.sheetnames

    fp = wb["_JSW_MRA_TEMPLATE_"]
    assert fp["A1"].value == "CUSTOMER_CODES_TEMPLATE"
    assert fp["A2"].value == "v1"


def test_template_headers_match_canonical_list() -> None:
    data = build_template_workbook()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb["Customer Codes"]
    headers = [cell.value for cell in ws[1]]
    assert headers == TEMPLATE_HEADERS
    assert len(headers) == 13


# ---------------------------------------------------------------------------
# Parser — happy path
# ---------------------------------------------------------------------------

def test_parse_accepts_official_template() -> None:
    data = build_template_workbook()
    rows, errors = parse_workbook(data)

    assert errors == []
    assert len(rows) == 1
    assert rows[0]["segment"] == "Retail"
    assert rows[0]["code"] == "40020365"
    assert rows[0]["ship_to_city"] == "Mumbai"
    assert rows[0]["rake"] == "RAKE-01"
    assert rows[0]["transport_mode"] == "RAKE"


def test_parse_accepts_template_with_multiple_data_rows() -> None:
    wb = openpyxl.load_workbook(io.BytesIO(build_template_workbook()), data_only=True)
    ws = wb["Customer Codes"]
    ws.append(["OEM", "8001", "Yard Internal", "Mumbai"] + [None] * 9)

    buf = io.BytesIO()
    wb.save(buf)
    rows, errors = parse_workbook(buf.getvalue())

    assert errors == []
    assert len(rows) == 2


# ---------------------------------------------------------------------------
# Parser — fingerprint enforcement
# ---------------------------------------------------------------------------

def test_parse_rejects_workbook_without_fingerprint() -> None:
    data = _build_workbook(
        TEMPLATE_HEADERS,
        [["Retail", "40020365", "Customer", "Mumbai"] + [None] * 9],
        include_fingerprint=False,
    )
    rows, errors = parse_workbook(data)

    assert rows == []
    assert len(errors) == 1
    assert errors[0].row == 0
    assert "official customer codes template" in errors[0].message.lower()


def test_parse_rejects_workbook_with_bad_fingerprint_value() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Codes"
    fp = wb.create_sheet(title="_JSW_MRA_TEMPLATE_")
    fp["A1"] = "WRONG_MARKER"
    fp["A2"] = "v1"
    ws.append(TEMPLATE_HEADERS)

    buf = io.BytesIO()
    wb.save(buf)
    rows, errors = parse_workbook(buf.getvalue())

    assert rows == []
    assert len(errors) == 1
    assert errors[0].row == 0


# ---------------------------------------------------------------------------
# Parser — header conformance
# ---------------------------------------------------------------------------

def test_parse_rejects_reordered_headers() -> None:
    reordered = TEMPLATE_HEADERS.copy()
    reordered[0], reordered[1] = reordered[1], reordered[0]
    data = _build_workbook(
        reordered,
        [["40020365", "Retail", "Customer", "Mumbai"] + [None] * 9],
    )
    rows, errors = parse_workbook(data)

    assert rows == []
    assert len(errors) == 1
    assert errors[0].row == 0
    assert "out of order" in errors[0].message.lower()


def test_parse_rejects_extra_column() -> None:
    headers = TEMPLATE_HEADERS + ["EXTRA"]
    data = _build_workbook(
        headers,
        [["Retail", "40020365", "Customer", "Mumbai"] + [None] * 10],
    )
    rows, errors = parse_workbook(data)

    assert rows == []
    assert len(errors) == 1
    assert errors[0].row == 0


def test_parse_accepts_missing_optional_column() -> None:
    headers = TEMPLATE_HEADERS[:-1]  # drop TRANSPORT MODE
    data = _build_workbook(
        headers,
        [["Retail", "40020365", "Customer", "Mumbai"] + [None] * 8],
    )
    rows, errors = parse_workbook(data)

    assert errors == []
    assert len(rows) == 1
    assert rows[0]["transport_mode"] is None


def test_parse_fills_missing_required_column_with_unknown() -> None:
    headers = [h for h in TEMPLATE_HEADERS if h != "Segment"]
    data = _build_workbook(
        headers,
        [["40020365", "Customer", "Mumbai"] + [None] * 9],
    )
    rows, errors = parse_workbook(data)

    assert errors == []
    assert len(rows) == 1
    assert rows[0]["segment"] == "unknown"


# ---------------------------------------------------------------------------
# Parser — row-level behaviour
# ---------------------------------------------------------------------------

def test_parse_skips_fully_empty_rows() -> None:
    data = _build_workbook(
        TEMPLATE_HEADERS,
        [
            ["Retail", "40020365", "Customer", "Mumbai"] + [None] * 9,
            [None] * 13,
            ["OEM", "8002", "Another", "Pune"] + [None] * 9,
        ],
    )
    rows, errors = parse_workbook(data)

    assert errors == []
    assert len(rows) == 2


def test_parse_fills_missing_required_field_with_unknown() -> None:
    data = _build_workbook(
        TEMPLATE_HEADERS,
        [
            ["Retail", "40020365", "", "Mumbai"] + [None] * 9,
        ],
    )
    rows, errors = parse_workbook(data)

    assert errors == []
    assert len(rows) == 1
    assert rows[0]["customer"] == "unknown"


# ---------------------------------------------------------------------------
# Parser — row limit guard
# ---------------------------------------------------------------------------

def test_parse_rejects_more_than_max_rows() -> None:
    from app.utils.customer_code.excel import MAX_IMPORT_ROWS

    data = _build_workbook(
        TEMPLATE_HEADERS,
        [["Retail", str(i), "Customer", "Mumbai"] + [None] * 9 for i in range(MAX_IMPORT_ROWS + 1)],
    )
    rows, errors = parse_workbook(data)

    assert rows == []
    assert len(errors) == 1
    assert errors[0].row == 0
    assert "row limit" in errors[0].message
