"""Unit tests for the combined report export (sheet selection + naming).

Mocks the report generator, stock/credit fetches and the rake drill-down so no
MongoDB connection is required.
"""

from __future__ import annotations

import asyncio
import io
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import load_workbook

from app.core.errors import ValidationError
from app.schemas.report import (
    CombinedExportBody,
    CombinedExportQuery,
    RakeDrilldownMergedRow,
    RakeDrilldownResponse,
    RakeDrilldownRow,
    RakeExclusion,
    ReportPivotRow,
    ReportResponse,
)
from app.services.report.export_combined import export_combined
from app.services.report.rake_drilldown import row_identity


def _fake_report() -> ReportResponse:
    row = ReportPivotRow(
        so_sales_org="1001", distr_chnl="OEM", sold_to_party="ACME",
        sales_office="Mumbai", party_code="40122581", ship_to_party="ACME ST",
        transport_mode="ROAD", destination="Pune", route="KAT1",
        rake_quantities={"KKU": 55.0}, total=55.0, nco_yes_do=0.0,
        nco_yes_do_count=0, blocked=False, credit_balance=5.0, required_credit=5.0,
        credit_sufficient=True, credit_status="", credit_note="Balance Available",
    )
    return ReportResponse(
        date="23-06-2026", report_type="jsw", region_id=None,
        region_name="All Regions", days_filter="include", ccas=["VJ0H"],
        has_stock=True, has_credit_report=True, coil_price_per_qty=1.0,
        rake_columns=["KKU"], rows=[row], grand_total=55.0, grand_nco_yes_do=0.0,
        grand_required_credit=5.0, rake_totals={"KKU": 55.0, "ROAD": 10.0},
        transport_mode_totals={"ROAD": 65.0},
    )


def _run(sheets: str, **patches) -> list[str]:
    """Run export_combined for *sheets* with patched deps; return sheet titles."""
    query = CombinedExportQuery(date="23-06-2026", report_type="jsw", sheets=sheets)
    cm = {
        "app.services.report.export_combined.generate_report": AsyncMock(return_value=_fake_report()),
        "app.services.report.export_combined.fetch_jsw_stock_docs": AsyncMock(return_value=[]),
        "app.services.report.export_combined.fetch_jvml_stock_docs": AsyncMock(return_value=[]),
        "app.services.report.export_combined.fetch_credit_report_docs": AsyncMock(return_value=[]),
        **patches,
    }
    stack = [patch(target, new=mock) for target, mock in cm.items()]
    for p in stack:
        p.start()
    try:
        data = asyncio.run(export_combined(query))
    finally:
        for p in stack:
            p.stop()
    return load_workbook(io.BytesIO(data)).sheetnames


def test_combined_selects_pivot_and_totals() -> None:
    names = _run("pivot,rake_totals")
    assert "BRANCH WISE PIVOT REPORT" in names
    assert "TOTAL RAKE REPORT" in names
    assert "Metadata" in names
    # Stock/credit not requested → absent.
    assert "JSW Stock" not in names and "Credit Report" not in names


def test_combined_selects_stock_and_credit() -> None:
    names = _run("jsw,credit")
    assert "JSW Stock" in names
    assert "Credit Report" in names
    assert "BRANCH WISE PIVOT REPORT" not in names


def test_combined_rake_breakdown_one_sheet_per_rake() -> None:
    """rake_merged + rake_unmerged → a sheet per unique rake, merged block first."""
    drill = RakeDrilldownResponse(
        rake="X", date="23-06-2026", region_id=None, region_name="All Regions",
        days_filter="include", rows=[], merged_rows=[], total_quantity=0.0,
    )
    names = _run(
        "rake_merged,rake_unmerged",
        **{
            "app.services.report.rake_breakdown_export._resolve_region_customers":
                AsyncMock(return_value=("All Regions", set(), {}, ["CWCJ", "ROAD"])),
            "app.services.report.rake_breakdown_export.rake_drilldown":
                AsyncMock(return_value=drill),
        },
    )
    assert names[:2] == ["CWCJ - Total Rake Wise", "ROAD - Total Rake Wise"]  # all merged first
    assert "CWCJ - Batch Rake Wise" in names and "ROAD - Batch Rake Wise" in names


def _drill_two_rows() -> RakeDrilldownResponse:
    """A KKU drill-down with two distinct jsw rows (30 + 25 = 55)."""
    common = dict(
        stock_type="jsw", so_sales_org="1001", distr_chnl="OEM", sold_to_party="ACME",
        sales_office="Mumbai", ship_to_party="ST", transport_mode="ROAD", destination="Pune",
    )
    row_a = RakeDrilldownRow(party_code="40000001", customer_name="ALPHA", stock_quantity=30.0, **common)
    row_b = RakeDrilldownRow(party_code="40000002", customer_name="BETA", stock_quantity=25.0, **common)
    mcommon = dict(
        so_sales_org="1001", distr_chnl="OEM", sales_office="Mumbai", sold_to_party="ACME",
        transport_mode="ROAD", destination="Pune",
    )
    m_a = RakeDrilldownMergedRow(party_code="40000001", customer_name="ALPHA", stock_quantity=30.0, **mcommon)
    m_b = RakeDrilldownMergedRow(party_code="40000002", customer_name="BETA", stock_quantity=25.0, **mcommon)
    return RakeDrilldownResponse(
        rake="KKU", date="23-06-2026", region_id=None, region_name="All Regions",
        days_filter="include", rows=[row_a, row_b], merged_rows=[m_a, m_b], total_quantity=55.0,
    )


def _report_two_parties() -> ReportResponse:
    """Pivot whose rows mirror the KKU drill rows (ALPHA 30 + BETA 25) plus an
    untouched CWCJ/RAIL row — so excluding ALPHA's identity nets it out of the
    recomputed pivot + rake + transport totals (the backend derives the subtraction
    from the keys, not the FE-sent ``subtract`` float)."""
    common = dict(
        so_sales_org="1001", distr_chnl="OEM", sold_to_party="ACME",
        sales_office="Mumbai", ship_to_party="ST", destination="Pune", route="KAT1",
        nco_yes_do=0.0, nco_yes_do_count=0, blocked=False, credit_balance=5.0,
        required_credit=5.0, credit_sufficient=True, credit_status="",
        credit_note="Balance Available",
    )
    alpha = ReportPivotRow(party_code="40000001", customer_name="ALPHA", transport_mode="ROAD",
                           rake_quantities={"KKU": 30.0}, total=30.0, **common)
    beta = ReportPivotRow(party_code="40000002", customer_name="BETA", transport_mode="ROAD",
                          rake_quantities={"KKU": 25.0}, total=25.0, **common)
    gamma = ReportPivotRow(party_code="40000003", customer_name="GAMMA", transport_mode="RAIL",
                           rake_quantities={"CWCJ": 10.0}, total=10.0, **common)
    return ReportResponse(
        date="23-06-2026", report_type="jsw", region_id=None, region_name="All Regions",
        days_filter="include", ccas=["VJ0H"], has_stock=True, has_credit_report=True,
        coil_price_per_qty=1.0, rake_columns=["KKU", "CWCJ"], rows=[alpha, beta, gamma],
        grand_total=65.0, grand_nco_yes_do=0.0, grand_required_credit=15.0,
        rake_totals={"KKU": 55.0, "CWCJ": 10.0},
        transport_mode_totals={"ROAD": 55.0, "RAIL": 10.0},
    )


def test_combined_exclusions_net_totals_pivot_and_breakdown() -> None:
    """An unchecked identity leaves the pivot, the rake/transport totals, and the
    breakdown sheets — all key-driven (the FE ``subtract`` float is ignored)."""
    drill = _drill_two_rows()
    excluded_key = row_identity(drill.rows[0])  # ALPHA / 30
    # subtract=0 on purpose: totals must drop from the KEYS, not the FE float.
    body = CombinedExportBody(exclusions={"KKU": RakeExclusion(keys=[excluded_key], subtract=0.0)})
    query = CombinedExportQuery(
        date="23-06-2026", report_type="jsw", sheets="pivot,rake_totals,rake_unmerged"
    )
    stack = [
        patch("app.services.report.export_combined.generate_report",
              new=AsyncMock(return_value=_report_two_parties())),
        patch("app.services.report.rake_breakdown_export._resolve_region_customers",
              new=AsyncMock(return_value=("All Regions", set(), {}, ["KKU"]))),
        patch("app.services.report.rake_breakdown_export.rake_drilldown",
              new=AsyncMock(return_value=drill)),
    ]
    for p in stack:
        p.start()
    try:
        data = asyncio.run(export_combined(query, body))
    finally:
        for p in stack:
            p.stop()

    wb = load_workbook(io.BytesIO(data))

    # TOTAL RAKE REPORT: KKU 55 → 25 (ALPHA out); CWCJ untouched at 10.
    totals = {r[0]: r[1] for r in wb["TOTAL RAKE REPORT"].iter_rows(values_only=True) if r and r[0]}
    assert totals.get("KKU") == 25.0
    assert totals.get("CWCJ") == 10.0

    # TRANSPORT MODE TOTAL: ROAD 55 → 25; RAIL untouched at 10.
    tm = {r[0]: r[1] for r in wb["TRANSPORT MODE TOTAL"].iter_rows(values_only=True) if r and r[0]}
    assert tm.get("ROAD") == 25.0
    assert tm.get("RAIL") == 10.0

    # BRANCH WISE PIVOT REPORT: ALPHA's party gone; BETA + GAMMA remain.
    pivot_cells = [c for row in wb["BRANCH WISE PIVOT REPORT"].iter_rows(values_only=True) for c in row if c is not None]
    assert "40000001" not in pivot_cells
    assert "40000002" in pivot_cells and "40000003" in pivot_cells

    # KKU - Batch Rake Wise: ALPHA (excluded) gone, BETA remains, Total recomputed to 25.
    cells = [c for row in wb["KKU - Batch Rake Wise"].iter_rows(values_only=True) for c in row if c is not None]
    assert "40000001" not in cells and "40000002" in cells
    assert any(c == "Total" for c in cells)
    assert 25.0 in cells and 55.0 not in cells


def test_combined_no_body_is_unchanged() -> None:
    """No exclusions body (GET path) leaves rake totals untouched."""
    drill = _drill_two_rows()
    query = CombinedExportQuery(date="23-06-2026", report_type="jsw", sheets="rake_totals")
    with patch("app.services.report.export_combined.generate_report",
               new=AsyncMock(return_value=_fake_report())):
        data = asyncio.run(export_combined(query))  # no body arg
    wb = load_workbook(io.BytesIO(data))
    totals = {r[0]: r[1] for r in wb["TOTAL RAKE REPORT"].iter_rows(values_only=True) if r and r[0]}
    assert totals.get("KKU") == 55.0 and totals.get("ROAD") == 10.0
    tm = {r[0]: r[1] for r in wb["TRANSPORT MODE TOTAL"].iter_rows(values_only=True) if r and r[0]}
    assert tm.get("ROAD") == 65.0  # transport mode untouched without a body


def test_combined_rejects_unknown_sheet_key() -> None:
    with pytest.raises(ValidationError):
        asyncio.run(
            export_combined(
                CombinedExportQuery(date="23-06-2026", report_type="jsw", sheets="bogus")
            )
        )


def test_combined_rejects_empty_sheet_selection() -> None:
    with pytest.raises(ValidationError):
        asyncio.run(
            export_combined(
                CombinedExportQuery(date="23-06-2026", report_type="jsw", sheets=" , ")
            )
        )
